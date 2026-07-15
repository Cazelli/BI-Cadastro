from __future__ import annotations

import argparse
from copy import copy
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Callable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
BASE_FILE = ROOT / "base_consolidada_copel.csv"
BASE_XLSX_FILE = ROOT / "base_consolidada_copel.xlsx"
ALERT_FILE = DATA_DIR / "ultima_atualizacao_alertas.csv"
SUMMARY_FILE = DATA_DIR / "ultima_atualizacao_resumo.json"
BACKUP_DIR = DATA_DIR / "backups"
REPORT_PATTERN = re.compile(
    r"^mdm-sandbox_clientes_novo-(\d{8})-(\d{8})\.csv$",
    re.IGNORECASE,
)
TITLE_CHANGE_START_DATE = pd.Timestamp("2026-03-01")
TRACKING_COLUMNS = ["DT_SITUACAO_UC", "DT_MUD_TIT", "MUD_TIT"]

ALERT_COLUMNS = [
    "NUM_UC",
    "GRUPO_UC",
    "TIPO_ALERTA",
    "CAMPO",
    "VALOR_ANTERIOR",
    "VALOR_NOVO",
    "DETALHE",
    "ARQUIVO_ORIGEM",
]
MONITORED_COLUMNS = [
    "SITUACAO_UC",
    "CLASSE",
    "GRUPO",
    "GD_BENE_INIC",
    "GD_BENE_FIM",
    "TIPO_GD_BENE",
    "DATA_INICIO_GD",
    "DATA_FIM_GD",
    "TIPO_GD_GERA",
    "TARIFA_SOCIAL",
    "TARIFA_BRANCA",
]


def text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    result = str(value).strip()
    return "" if result.lower() in {"nan", "none", "<na>"} else result


def upper(value: object) -> str:
    return text(value).upper()


def identifier(value: object) -> str:
    result = text(value)
    if not result:
        return ""
    if re.fullmatch(r"\d+(?:\.0+)?", result):
        return str(int(float(result)))
    return result


def integer(value: object) -> str:
    result = text(value)
    if not result:
        return ""
    try:
        return str(int(float(result.replace(",", "."))))
    except ValueError:
        return result


def iso_date(value: object) -> str:
    result = text(value)
    if not result:
        return ""
    parsed = pd.to_datetime(result, errors="coerce", dayfirst="/" in result)
    if pd.isna(parsed):
        raise ValueError(f"Data inválida no relatório: {result!r}")
    return parsed.strftime("%Y-%m-%d")


def yes_no(value: object) -> str:
    normalized = upper(value)
    if normalized in {"S", "SIM", "GERADORA", "TRUE", "1"}:
        return "S"
    if normalized in {"N", "NAO", "NÃO", "FALSE", "0", ""}:
        return "N"
    return normalized


def social_tariff(value: object) -> str:
    normalized = upper(value)
    return "N" if normalized in {"", "N", "NAO", "NÃO", "0"} else "S"


def group_label(value: object) -> str:
    normalized = upper(value)
    return {
        "ATIVO": "Tratamento",
        "ATIVOS": "Tratamento",
        "ATIVA": "Tratamento",
        "ATIVAS": "Tratamento",
        "CONTROLE": "Controle",
        "RESERVA": "Reserva",
    }.get(normalized, text(value) or "Não informado")


Normalizer = Callable[[object], str]

# Only fields with a direct, reliable correspondence are updated. Project fields
# such as ETAPA, DT_ATIVACAO and FINALIDADE remain under the consolidated base's
# existing governance.
FIELD_MAP: dict[str, tuple[str, Normalizer]] = {
    "uc_aneel": ("NUM_UC_ANEEL", identifier),
    "cliente": ("COD_CLIENTE", identifier),
    "nome": ("NOME_TITULAR", text),
    "celular": ("CELULAR", identifier),
    "email": ("EMAIL", text),
    "medidor": ("NIO", identifier),
    "sub_grupo": ("CLASSE", upper),
    "classe_principal": ("GRUPO", integer),
    "situacao_uc": ("SITUACAO_UC", upper),
    "tipo_fase": ("TIPO_FASE", upper),
    "municipio": ("LOCAL", upper),
    "data_situacao": ("DT_SITUACAO_UC", iso_date),
    "max_data_tt": ("DT_MUD_TIT", iso_date),
    "inicio_beneficiaria": ("GD_BENE_INIC", iso_date),
    "fim_benificiaria": ("GD_BENE_FIM", iso_date),
    # TIPO_GD_BENE is not modalidade_geracao. This optional source column will
    # only be used once it is delivered explicitly in a future report.
    "tipo_gd_bene": ("TIPO_GD_BENE", upper),
    "data_inicio_gd": ("DATA_INICIO_GD", iso_date),
    "data_fim_gd": ("DATA_FIM_GD", iso_date),
    "tipo_gd": ("TIPO_GD_GERA", upper),
    "geracao_propria": ("POSSUI_GD_CLIENTE", yes_no),
    "tarifa_branca": ("TARIFA_BRANCA", yes_no),
    "baixa_renda": ("TARIFA_SOCIAL", social_tariff),
}
OPTIONAL_SOURCE_COLUMNS = {"tipo_gd_bene"}
TARGET_NORMALIZERS: dict[str, Normalizer] = {
    target: (
        upper
        if target in {"POSSUI_GD_CLIENTE", "TARIFA_SOCIAL", "TARIFA_BRANCA"}
        else normalizer
    )
    for target, normalizer in FIELD_MAP.values()
}
TARGET_NORMALIZERS["MUD_TIT"] = upper


def newest_report() -> Path:
    candidates: list[tuple[str, str, Path]] = []
    for path in DATA_DIR.glob("mdm-sandbox_clientes_novo-*.csv"):
        match = REPORT_PATTERN.fullmatch(path.name)
        if match:
            datetime.strptime(match.group(1), "%Y%m%d")
            datetime.strptime(match.group(2), "%Y%m%d")
            candidates.append((match.group(2), match.group(1), path))
    if not candidates:
        raise FileNotFoundError(
            "Nenhum relatório data/mdm-sandbox_clientes_novo-YYYYMMDD-YYYYMMDD.csv encontrado."
        )
    return max(candidates)[2]


def read_summary() -> dict:
    if not SUMMARY_FILE.exists():
        return {}
    return json.loads(SUMMARY_FILE.read_text(encoding="utf-8"))


def validate_columns(frame: pd.DataFrame, required: set[str], label: str) -> None:
    missing = sorted(required.difference(frame.columns))
    if missing:
        raise ValueError(f"Colunas ausentes em {label}: {', '.join(missing)}")


def atomic_csv(frame: pd.DataFrame, destination: Path) -> None:
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    frame.to_csv(temporary, index=False, encoding="utf-8-sig", lineterminator="\n")
    temporary.replace(destination)


def atomic_json(payload: dict, destination: Path) -> None:
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(destination)


def update_excel_tracking_columns(base: pd.DataFrame, report: Path) -> None:
    if not BASE_XLSX_FILE.exists():
        return

    excel_backup = BACKUP_DIR / f"base_consolidada_xlsx_before_{report.stem}.xlsx"
    if not excel_backup.exists():
        shutil.copy2(BASE_XLSX_FILE, excel_backup)

    workbook = load_workbook(BASE_XLSX_FILE)
    worksheet = (
        workbook["base_consolidada"]
        if "base_consolidada" in workbook.sheetnames
        else workbook.active
    )
    headers = {
        text(cell.value): cell.column
        for cell in worksheet[1]
        if text(cell.value)
    }
    if "NUM_UC" not in headers:
        raise ValueError("A planilha base_consolidada_copel.xlsx não possui NUM_UC.")

    for column in TRACKING_COLUMNS:
        if column in headers:
            continue
        new_column = worksheet.max_column + 1
        header_cell = worksheet.cell(row=1, column=new_column, value=column)
        previous_header = worksheet.cell(row=1, column=new_column - 1)
        if previous_header.has_style:
            header_cell._style = copy(previous_header._style)
            header_cell.font = copy(previous_header.font)
            header_cell.fill = copy(previous_header.fill)
            header_cell.border = copy(previous_header.border)
            header_cell.alignment = copy(previous_header.alignment)
            header_cell.number_format = previous_header.number_format
            header_cell.protection = copy(previous_header.protection)
        headers[column] = new_column

    if worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )

    base_by_uc = base.set_index("NUM_UC", drop=False)
    uc_column = headers["NUM_UC"]
    for row_number in range(2, worksheet.max_row + 1):
        uc = identifier(worksheet.cell(row=row_number, column=uc_column).value)
        if not uc or uc not in base_by_uc.index:
            continue
        for column in TRACKING_COLUMNS:
            value = text(base_by_uc.at[uc, column])
            cell = worksheet.cell(row=row_number, column=headers[column])
            if column.startswith("DT_") and value:
                cell.value = datetime.strptime(value, "%Y-%m-%d").date()
                cell.number_format = "yyyy-mm-dd"
            else:
                cell.value = value or None

    temporary = BASE_XLSX_FILE.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(BASE_XLSX_FILE)


def add_alert(
    alerts: list[dict[str, str]],
    uc: str,
    group: str,
    alert_type: str,
    field: str,
    previous: str,
    new: str,
    detail: str,
    source: str,
) -> None:
    alerts.append(
        {
            "NUM_UC": uc,
            "GRUPO_UC": group,
            "TIPO_ALERTA": alert_type,
            "CAMPO": field,
            "VALOR_ANTERIOR": previous,
            "VALOR_NOVO": new,
            "DETALHE": detail,
            "ARQUIVO_ORIGEM": source,
        }
    )


def change_detail(field: str, new: str) -> tuple[str, str]:
    if field == "SITUACAO_UC" and new == "DS":
        return "Desligamento", "A situação da UC mudou para desligada (DS)."
    if field == "CLASSE":
        if new != "B1":
            return "Mudança de Classe", "A classe esperada é B1."
        return "Alteração cadastral", "A classe da UC foi alterada."
    if field == "GRUPO":
        if new != "1":
            return "Mudança de Classe", "O grupo esperado é 1."
        return "Alteração cadastral", "O grupo da UC foi alterado."
    if field in {"TARIFA_SOCIAL", "TARIFA_BRANCA"} and new == "S":
        return "Tarifa Especial Ativada", f"{field} passou a ser S."
    if field.startswith("GD_") or field in {
        "TIPO_GD_BENE",
        "DATA_INICIO_GD",
        "DATA_FIM_GD",
        "TIPO_GD_GERA",
    }:
        return "Alteração GD", "Informação de geração distribuída alterada."
    return "Alteração cadastral", "Valor cadastral alterado no último relatório."


def process_report(report: Path, force: bool = False) -> dict:
    if not BASE_FILE.exists():
        raise FileNotFoundError(f"Base consolidada não encontrada: {BASE_FILE}")

    previous_summary = read_summary()
    if previous_summary.get("arquivo_origem") == report.name and not force:
        print(f"Relatório já processado: {report.name}")
        return previous_summary

    base = pd.read_csv(BASE_FILE, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    incoming = pd.read_csv(
        report,
        sep=";",
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    )
    base.columns = [text(column) for column in base.columns]
    incoming.columns = [text(column).lower() for column in incoming.columns]

    for column in TRACKING_COLUMNS:
        if column not in base.columns:
            base[column] = ""
    validate_columns(base, {"NUM_UC", "SITUACAO_INICIAL", *TARGET_NORMALIZERS}, "base")
    required_sources = set(FIELD_MAP).difference(OPTIONAL_SOURCE_COLUMNS)
    validate_columns(incoming, {"uc", *required_sources}, report.name)

    base["NUM_UC"] = base["NUM_UC"].map(identifier)
    incoming["uc"] = incoming["uc"].map(identifier)
    if base["NUM_UC"].eq("").any() or incoming["uc"].eq("").any():
        raise ValueError("Existem UCs sem identificador na base ou no relatório.")
    if base["NUM_UC"].duplicated().any():
        duplicates = base.loc[base["NUM_UC"].duplicated(False), "NUM_UC"].unique()
        raise ValueError(f"UCs duplicadas na base: {', '.join(duplicates[:10])}")
    if incoming["uc"].duplicated().any():
        duplicates = incoming.loc[incoming["uc"].duplicated(False), "uc"].unique()
        raise ValueError(f"UCs duplicadas no relatório: {', '.join(duplicates[:10])}")

    base_index = {uc: index for index, uc in base["NUM_UC"].items()}
    alerts: list[dict[str, str]] = []
    changed_ucs: set[str] = set()
    ignored_extra_ucs = 0
    matched_ucs = 0
    matched_uc_ids: set[str] = set()

    for _, source_row in incoming.iterrows():
        uc = source_row["uc"]
        if uc not in base_index:
            ignored_extra_ucs += 1
            continue

        converted = {
            target: normalizer(source_row[source])
            for source, (target, normalizer) in FIELD_MAP.items()
            if source in incoming.columns
        }
        title_change_date = converted.get("DT_MUD_TIT", "")
        converted["MUD_TIT"] = (
            "S"
            if title_change_date
            and pd.Timestamp(title_change_date) >= TITLE_CHANGE_START_DATE
            else ""
        )
        matched_ucs += 1
        matched_uc_ids.add(uc)
        index = base_index[uc]
        group = group_label(base.at[index, "SITUACAO_INICIAL"])
        for field in MONITORED_COLUMNS:
            if field not in converted:
                continue
            normalizer = TARGET_NORMALIZERS[field]
            previous = normalizer(base.at[index, field])
            new = converted[field]
            if previous != new:
                if field in {"TARIFA_SOCIAL", "TARIFA_BRANCA"} and new != "S":
                    continue
                alert_type, detail = change_detail(field, new)
                add_alert(
                    alerts,
                    uc,
                    group,
                    alert_type,
                    field,
                    previous,
                    new,
                    detail,
                    report.name,
                )
                changed_ucs.add(uc)

        # Keep all mapped consolidated fields current, including non-alert fields.
        for field, new in converted.items():
            old = TARGET_NORMALIZERS[field](base.at[index, field])
            if old != new:
                base.at[index, field] = new

        # Class/group are persistent compliance rules, so an unchanged invalid
        # value remains visible in every report that contains the UC.
        for field, expected in (("CLASSE", "B1"), ("GRUPO", "1")):
            current = converted[field]
            already_alerted = any(
                alert["NUM_UC"] == uc and alert["CAMPO"] == field
                for alert in alerts
            )
            if current != expected and not already_alerted:
                add_alert(
                    alerts,
                    uc,
                    group,
                    "Mudança de Classe",
                    field,
                    current,
                    current,
                    f"Valor esperado: {expected}.",
                    report.name,
                )

    missing_update_ucs = set(base_index).difference(matched_uc_ids)
    for uc in sorted(missing_update_ucs, key=lambda value: (len(value), value)):
        index = base_index[uc]
        add_alert(
            alerts,
            uc,
            group_label(base.at[index, "SITUACAO_INICIAL"]),
            "Sem atualização",
            "PRESENCA_NO_RELATORIO",
            "Presente na base",
            "Ausente no relatório",
            "A UC não recebeu dados no último relatório processado.",
            report.name,
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"base_consolidada_before_{report.stem}.csv"
    if not backup.exists():
        shutil.copy2(BASE_FILE, backup)

    alerts_frame = pd.DataFrame(alerts, columns=ALERT_COLUMNS)
    atomic_csv(base, BASE_FILE)
    atomic_csv(alerts_frame, ALERT_FILE)
    update_excel_tracking_columns(base, report)

    filename_match = REPORT_PATTERN.fullmatch(report.name)
    summary = {
        "arquivo_origem": report.name,
        "periodo_inicio": filename_match.group(1) if filename_match else "",
        "periodo_fim": filename_match.group(2) if filename_match else "",
        "processado_em": datetime.now().astimezone().isoformat(timespec="seconds"),
        "ucs_no_relatorio": int(len(incoming)),
        "ucs_correspondentes": int(matched_ucs),
        "ucs_extras_ignoradas": int(ignored_extra_ucs),
        "ucs_sem_atualizacao": int(len(missing_update_ucs)),
        "ucs_alteradas": int(len(changed_ucs)),
        "alertas": int(len(alerts_frame)),
        "total_base": int(len(base)),
    }
    atomic_json(summary, SUMMARY_FILE)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Atualiza a base consolidada e gera alertas do relatório mais recente."
    )
    parser.add_argument(
        "--input",
        type=Path,
        help="Relatório específico. Por padrão, usa o arquivo datado mais recente em data/.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Reprocessa o relatório mesmo se ele já constar como a última atualização.",
    )
    args = parser.parse_args()
    report = args.input.resolve() if args.input else newest_report()
    if not report.exists() or not REPORT_PATTERN.fullmatch(report.name):
        raise ValueError(
            "O relatório deve existir e seguir o nome "
            "mdm-sandbox_clientes_novo-YYYYMMDD-YYYYMMDD.csv."
        )
    summary = process_report(report, force=args.force)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
